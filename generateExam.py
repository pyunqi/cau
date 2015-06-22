#-*- coding: utf-8 -*-
import uniout
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

filename = os.path.expanduser('res.xlsx')
try:
    os.remove(filename)
except OSError:
    pass

orgwb = load_workbook('org.xlsx',use_iterators=True) # 数据源

tarwb = Workbook()
dest_filename = 'res.xlsx'

ws1 = tarwb.active
ws1.title = "MT"
ws2 = tarwb.create_sheet(title="STPT")
ws3 = tarwb.create_sheet(title="STDS")
ws4 = tarwb.create_sheet(title="DT")
ws5 = tarwb.create_sheet(title="SA")

# print orgwb.get_sheet_names()
ws  = orgwb.get_sheet_by_name('1')

# 5个sheet 分别记录正在写入的行数
sa = 0
mt = 0
stpt = 0
dt = 0
stds = 0

# 读一行 写一行
for row in ws.iter_rows():
    
    if type(row[8].value) is not unicode:
        answers = str(row[8].value)
    else:
        answers = row[8].value.encode('utf-8')

    qtype_1 = row[2].value
    qtype_2 = row[3].value
    qtype_3 = row[4].value
    qtype_4 = row[5].value
    qtype_5 = row[6].value
    question_num = row[0].value

    (one,two,three,four,five) = ('N','N','N','N','N')

    for answer in answers.split('、'):
        if (answer == '1'):
            one = 'Y'
        if (answer == '2'):
            two = 'Y'
        if (answer == '3'):
            three = 'Y' 
        if (answer == '4'):
            four = 'Y'
        if (answer == '5'):
            five = 'Y'  
    # print qtype_1,qtype_2,qtype_3,qtype_4,qtype_5
    if (qtype_1 == 1):
        mt = mt+1
        ws1.cell(column=1, row=mt).value = question_num
        ws1.cell(column=2, row=mt, value= row[10].value)
        ws1.cell(column=3, row=mt, value= row[11].value)
        ws1.cell(column=4, row=mt).value= one
        ws1.cell(column=5, row=mt, value= row[12].value)
        ws1.cell(column=6, row=mt).value= two
        ws1.cell(column=7, row=mt, value= row[13].value)
        ws1.cell(column=8, row=mt).value= three
        ws1.cell(column=9, row=mt, value= row[14].value)
        ws1.cell(column=10, row=mt).value= four
        ws1.cell(column=11, row=mt, value= row[15].value)
        ws1.cell(column=12, row=mt).value= five
    if (qtype_2 == 1):
        stpt = stpt + 1
        ws2.cell(column=1, row=stpt).value = question_num
        ws2.cell(column=2, row=stpt, value= row[10].value)
        ws2.cell(column=3, row=stpt, value= row[11].value)
        ws2.cell(column=4, row=stpt).value= one
        ws2.cell(column=5, row=stpt, value= row[12].value)
        ws2.cell(column=6, row=stpt).value= two
        ws2.cell(column=7, row=stpt, value= row[13].value)
        ws2.cell(column=8, row=stpt).value= three
        ws2.cell(column=9, row=stpt, value= row[14].value)
        ws2.cell(column=10, row=stpt).value= four
        ws2.cell(column=11, row=stpt, value= row[15].value)
        ws2.cell(column=12, row=stpt).value= five
    if (qtype_3 == 1):
        stds = stds + 1
        ws3.cell(column=1, row=stds).value  = question_num
        ws3.cell(column=2, row=stds, value= row[10].value)
        ws3.cell(column=3, row=stds, value= row[11].value)
        ws3.cell(column=4, row=stds).value= one
        ws3.cell(column=5, row=stds, value= row[12].value)
        ws3.cell(column=6, row=stds).value = two
        ws3.cell(column=7, row=stds, value= row[13].value)
        ws3.cell(column=8, row=stds).value= three
        ws3.cell(column=9, row=stds, value= row[14].value)
        ws3.cell(column=10, row=stds).value= four
        ws3.cell(column=11, row=stds, value= row[15].value)
        ws3.cell(column=12, row=stds).value= five
    if (qtype_4 == 1):
        dt = dt + 1
        ws4.cell(column=1, row=dt).value = question_num
        ws4.cell(column=2, row=dt, value= row[10].value)
        ws4.cell(column=3, row=dt, value= row[11].value)
        ws4.cell(column=4, row=dt).value= one
        ws4.cell(column=5, row=dt, value= row[12].value)
        ws4.cell(column=6, row=dt).value= two
        ws4.cell(column=7, row=dt, value= row[13].value)
        ws4.cell(column=8, row=dt).value= three
        ws4.cell(column=9, row=dt, value= row[14].value)
        ws4.cell(column=10, row=dt).value= four
        ws4.cell(column=11, row=dt, value= row[15].value)
        ws4.cell(column=12, row=dt).value= five
    if (qtype_5 == 1):
        sa = sa + 1
        ws5.cell(column=1, row=sa).value = question_num
        ws5.cell(column=2, row=sa, value= row[10].value)
        ws5.cell(column=3, row=sa, value= row[11].value)
        ws5.cell(column=4, row=sa).value= one
        ws5.cell(column=5, row=sa, value= row[12].value)
        ws5.cell(column=6, row=sa).value = two
        ws5.cell(column=7, row=sa, value= row[13].value)
        ws5.cell(column=8, row=sa).value= three
        ws5.cell(column=9, row=sa, value= row[14].value)
        ws5.cell(column=10, row=sa).value= four
        ws5.cell(column=11, row=sa, value= row[15].value)
        ws5.cell(column=12, row=sa).value= five

tarwb.save(filename = dest_filename)

    # data = {
    #     'type':  row[2].value, # Column I  选项答案
    #     'answers':  row[8].value, # Column I  选项答案
    #     'question': row[10].value, # Column K ，问题
    #     'an_one':  row[11].value, # Column L 选项 1
    #     'an_two':  row[12].value, # Column M 选项 2
    #     'an_three':  row[13].value, # Column N  选项 3
    #     'an_four':  row[14].value, # Column O 选项 4
    #     'an_five':  row[15].value, # Column P 选项 5
    # }



    # print  (data['answers'].encode('utf-8'), '::', data['question'].encode('utf-8'), '::', 
    #         data['an_one'].encode('utf-8'),data['an_two'].encode('utf-8'),data['an_three'].encode('utf-8'), 
    #         data['an_four'].encode('utf-8'), data['an_five'])

